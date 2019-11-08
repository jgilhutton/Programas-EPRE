from os import walk, mkdir, getcwd, rename
from re import search
from time import strptime
import openpyxl
import xlrd
import zipfile

meses = {'1':'Enero','2':'Febrero','3':'Marzo','4':'Abril','5':'Mayo','6':'Junio','7':'Julio','8':'Agosto','9':'Septiembre','O':'Octubre','N':'Noviembre','D':'Diciembre','10':'Octubre','11':'Noviembre','12':'Diciembre'}
mediciones = []
remediciones = []
carpetas = []
resultadosMediciones = []
sheetName = 'Rdo Mediciones Usuario-Centros'


def getFiles(directory):
 files = []
 for root, dir, file in walk(directory):
  files.append(file)
 files = files[0]
 return files

def getFecha(file):
 dia = file[2:4]
 mes = meses[file[4]]
 anio = file[5]
 cola = file[6:]
 return '{} de {}'.format(dia,mes)

def fechaRemedicion(file):
	if tipoExcel == 'xlsx':
		for i in range(9,1000):
			if resultadosMediciones['Q{}'.format(i)].value == file:
				fecha = resultadosMediciones['D{}'.format(i)].value
				break
	elif tipoExcel == 'xls':
		for i in range(9,1000):
			if resultadosMediciones.cell(i,16).value == file:    
				fecha = resultadosMediciones.cell(i,3).value
				break
	if fecha:
		fecha = strptime(fecha,'%d/%m/%Y %H:%M:%S')
		nombreCarpeta = '{} de {}'.format(str(fecha.tm_mday).zfill(2), meses[str(fecha.tm_mon)])
		return nombreCarpeta
	else: return False

def getExcel(nameList):
 tipo = ''
 for i in nameList:
  if i[-5:] == '.xlsx':
   resultadosMediciones = openpyxl.load_workbook('{}/{}'.format(cwd,i))[sheetName]
   tipo = 'xlsx'
   break
  elif i[-4:] == '.xls':
   wb = xlrd.open_workbook('{}/{}'.format(cwd,i))
   resultadosMediciones = wb.sheet_by_name(sheetName)
   tipo = 'xls'
   break
 return [resultadosMediciones,tipo]
    
if __name__ == '__main__':
	cwd = input('Carpeta:> ')
	if not cwd:
		cwd = getcwd()

	files = getFiles(cwd)
	zips = [x for x in files if x[-4:] == '.zip']
	if not zips: exit()
	if len(zips) == 1:
	 index = 0
	else:
	 for i in range(len(zips)):
	  print('{}) {}'.format(i+1,zips[i]))
	 index = int(input('Ingrese numero de comprimido> '))-1
	zip_ref = zipfile.ZipFile('{}/{}'.format(cwd,zips[index]), 'r')
	zip_ref.extractall(cwd)
	resultadosMediciones, tipoExcel = getExcel(zip_ref.namelist())
	zip_ref.close()

	files = getFiles(cwd)
	for file in files:
	 if search('\d{4}[\d,O,N,D][\d]O.+',file):
	  mediciones.append(file)
	 elif search('\d{4}[\d,O,N,D][\d][F,R].+',file):
	  remediciones.append(file)

	for medicion in mediciones:
	 carpeta = getFecha(medicion)
	 carpetas.append(carpeta)
	carpetas = list(set(carpetas))
	for carpeta in carpetas:
	 try:
	  mkdir('{}/{}'.format(cwd,carpeta))
	 except:
	  continue

	for file in mediciones:
	 origen = '{}\{}'.format(cwd,file)
	 destino = '{}\{}\{}'.format(cwd,getFecha(file),file)
	 rename(origen,destino)

	for file in remediciones:
		origen = '{}\{}'.format(cwd,file)
		fecha = ''
		try:
			fecha = fechaRemedicion(file)
		except: pass
		if not fecha: continue

		try:
			mkdir('{}/{}'.format(cwd,fecha))
		except: pass
		try:
			mkdir('{}/{}/remediciones'.format(cwd,fecha))
		except: pass

		destino = '{}/{}/remediciones/{}'.format(cwd,fecha,file)
		rename(origen,destino)