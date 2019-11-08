import pyodbc
import pickle
import openpyxl
from re import search
from os.path import isfile

motivosExcluidos = ['CSC','FIU','FCL','FOT','FEX']
# motivosExcluidos = ['CSC','FIU']

dbInterrupciones = 'D:/Documents and Settings/Pasante/Escritorio/Programas Juani/CCSPU y TCSPU/Anexo XX TABLAS EPRE 1802  definitiva 26-08-2019.mdb'
INTpickle = 'CCSPU y TCSPU/INTpickle'
tablaUsuarios = 'datos_comerciales_del_usuario'
tablaSIDAC = 'suministros_sidac'
tablaIUM = 'Interrupciones_por_usuario_Mensual'
tablaIRM = 'Interrupciones_Reposiciones_Mensual'
plantilla = 'Recursos/Indices.xlsx'

nivelesCalidad = {1:{'CCSPU':0,'TCSPU':0,'Factor':0.4},
			2:{'CCSPU':3.5,'TCSPU':290,'Factor':0.55},
			3:{'CCSPU':4.1,'TCSPU':360,'Factor':0.7},
			4:{'CCSPU':4.7,'TCSPU':430,'Factor':0.85},
			5:{'CCSPU':5.3,'TCSPU':500,'Factor':1},
		}
############################################################################################
############################################################################################
############################################################################################
############################################################################################

def salir(mensaje):
	print(mensaje)
	exit()
	
def conectarConDB():
	try:
		conexion = pyodbc.connect(r'Driver={Microsoft Access Driver (*.mdb, *.accdb)};DBQ=%s;' %dbInterrupciones)
		return conexion.cursor()
	except Exception as e:
		salir('Hubo un error al intentar conectarse con la base de datos "{}":\n{}'.format(dbInterrupciones,e))
		return None

def getCantidadUsuarios(cursor):
	query = 'SELECT COUNT(Suministro) FROM [{}];'.format(tablaUsuarios)
	cursor.execute(query)
	Q = int(cursor.fetchall()[0][0])
	query = 'SELECT COUNT(CLIENTE) FROM [{}] WHERE {}.ESTADO = \'1\';'.format(tablaSIDAC,tablaSIDAC)
	cursor.execute(query)
	QSS = int(cursor.fetchall()[0][0])
	return Q,QSS
	
def calcularNivelCalidad(CCSPU,TCSPU):
	niveles = sorted(nivelesCalidad.keys())
	nivelC = 0
	nivelT = 0
	for nivel in niveles:
		if CCSPU > nivelesCalidad[nivel]['CCSPU']: nivelC = nivel
		if TCSPU > nivelesCalidad[nivel]['TCSPU']: nivelT = nivel
	nivel = max((nivelC,nivelT))
	return nivel,nivelesCalidad[nivel]['Factor']
		
def getInterrupciones(cursor):
	query='SELECT DISTINCT [IUM].Interrupcion, [IUM].Orden_Reposicion, C.cuenta, [IUM].Inicio, [IUM].Final,[IRM].Motivo_EPRE FROM ((SELECT Interrupcion & Orden_Reposicion AS corte,COUNT(Interrupcion & Orden_Reposicion) AS cuenta FROM [IUM] GROUP BY Interrupcion & Orden_Reposicion) AS C INNER JOIN [IUM] ON [IUM].Interrupcion & [IUM].Orden_Reposicion = C.corte) INNER JOIN [IRM] ON ([IRM].Interrupcion = [IUM].Interrupcion AND [IRM].Orden_Reposicion = [IUM].Orden_Reposicion);'
	query = query.replace('IRM',tablaIRM)
	query = query.replace('IUM',tablaIUM)

	try: cursor.execute(query)
	except Exception as e:
		salir('ERROR\nHubo un error al intentar obtener las interrupciones\n{}'.format(e))
	return cursor
	
def fillColor():
	fill = openpyxl.styles.PatternFill(start_color='FFAAAAAA',
                   end_color='FFAAAAAA',
                   fill_type='solid')
	return fill	
	
def getDatos(interrupciones,totalUsuarios,totalUsuariosSIDAC):
	totalUsuariosAfectados = sum([x[2] for x in interrupciones])
	totalMinutosDeCorte = sum([int(x[2])*(x[4]-x[3]).total_seconds()/60 for x in interrupciones])
	CCSPU = totalUsuariosAfectados/totalUsuarios
	CCSPU_sidac = totalUsuariosAfectados/totalUsuariosSIDAC
	sumatoriaUporT = totalMinutosDeCorte
	TCSPU = sumatoriaUporT/totalUsuarios
	TCSPU_sidac = sumatoriaUporT/totalUsuariosSIDAC
	nivel,factor = calcularNivelCalidad(CCSPU,TCSPU)
	nivelSIDAC,factorSIDAC = calcularNivelCalidad(CCSPU_sidac,TCSPU_sidac)
	return {'totalUsuariosAfectados':totalUsuariosAfectados, 'totalMinutosDeCorte':totalMinutosDeCorte, 'CCSPU':CCSPU, 'CCSPU_sidac':CCSPU_sidac, 'sumatoriaUporT':sumatoriaUporT, 'TCSPU':TCSPU, 'TCSPU_sidac':TCSPU_sidac, 'nivel':nivel, 'factor':factor, 'nivelSIDAC':nivelSIDAC, 'factorSIDAC':factorSIDAC}
	
def completar(datos,fila,mes='Semestre'):
	ws['A{}'.format(fila+1)].value = 'Motivo'
	ws['B{}'.format(fila+1)].value = 'CCSPU'
	ws['C{}'.format(fila+1)].value = 'TCSPU minutos'
	ws['D{}'.format(fila+1)].value = 'TCSPU horas'
	ws['E{}'.format(fila+1)].value = 'Cant. Usuarios'
	ws['F{}'.format(fila+1)].value = 'Factor de Estimulo'
	ws['G{}'.format(fila+1)].value = 'Nivel'
	ws['A{}'.format(fila+2)].value = 'Calculado con datos comerciales (Activos)'
	ws['A{}'.format(fila+3)].value = 'Calculado con datos suministro sidac'
	ws['B{}'.format(fila+2)].value = datos['CCSPU']
	ws['B{}'.format(fila+3)].value = datos['CCSPU_sidac']
	ws['C{}'.format(fila+2)].value = datos['TCSPU']
	ws['C{}'.format(fila+3)].value = datos['TCSPU_sidac']
	ws['D{}'.format(fila+2)].value = datos['TCSPU']/60
	ws['D{}'.format(fila+3)].value = datos['TCSPU_sidac']/60
	ws['E{}'.format(fila+2)].value = totalUsuarios
	ws['E{}'.format(fila+3)].value = totalUsuariosSIDAC
	ws['F{}'.format(fila+2)].value = datos['factor']
	ws['F{}'.format(fila+3)].value = datos['factorSIDAC']
	ws['G{}'.format(fila+2)].value = datos['nivel']
	ws['G{}'.format(fila+3)].value = datos['nivelSIDAC']
	fila += 5
	return fila

if __name__ == '__main__':
	dbInterrupciones = dbInterrupciones.replace('\\','/')
	mesesDict = {1:'Enero',2:'Febrero',3:'Marzo',4:'Abril',5:'Mayo',6:'Junio',7:'Julio',8:'Agosto',9:'Septiembre',10:'Octubre',11:'Noviembre',12:'Diciembre'}
	print(
	"""
	1) Verifica la existencia de la base de datos:
	"{}"
	2) Dentro de la base de datos tienen que estar las siguientes tablas:
		a) IRM:      {}
		b) IUM:      {}
		c) SIDAC:    {}
		d) USUARIOS: {}
	3) Se excluyen los siguientes motivos:
		{}
		
	Enter para continuar
	""".format(dbInterrupciones,tablaIRM,tablaIUM,tablaSIDAC,tablaUsuarios,','.join(motivosExcluidos))
	)
	input()

	if not isfile(INTpickle):
		print('Generando pickle para interrupciones...\r',end='Generando pickle para interrupciones... ')
		ints = getInterrupciones(conectarConDB())
		interrupciones = ints.fetchall()
		with open(INTpickle,'wb') as ium:
			pickle.dump(interrupciones,ium)
		print('OK')
	else:
		print('Usando pickle previo para IUM')
		with open(INTpickle,'rb') as ium:
			try:
				interrupciones = pickle.load(ium)
			except EOFError:
				remove(INTpickle)
				salir('Quedó un pickle viejo corrupto. Ya lo eliminé. Echá a andar el programa devuelta')		
				
	motivosExcluidos = '|'.join(motivosExcluidos)
	interrupciones = list(filter(lambda x:not search(motivosExcluidos,x[-1]),interrupciones))
	interrupciones = list(filter(lambda x:((x[4]-x[3]).total_seconds()/60) > 3,interrupciones))

	wb = openpyxl.load_workbook(plantilla)
	ws = wb['Indice Semestral']
	totalUsuarios,totalUsuariosSIDAC = getCantidadUsuarios(conectarConDB())
	datosSemestre = getDatos(interrupciones,totalUsuarios,totalUsuariosSIDAC)
	ws['B1'].value = datosSemestre['totalUsuariosAfectados']
	ws['B2'].value = datosSemestre['totalMinutosDeCorte']
	ws['B4'].value = totalUsuarios
	ws['B5'].value = totalUsuariosSIDAC
	fila = 7
	fila = completar(datosSemestre,fila)
	fila+=1
	print('Generando plantilla excel... \r',end='Generando plantilla excel... ')

	meses = sorted(mesesDict)
	for mes in meses:
		cortes = list(filter(lambda x: x,filter(lambda x: x[3].month == mes,interrupciones)))
		ws['A{}'.format(fila)].value = mesesDict[mes]
		if cortes:
			datosMes = getDatos(cortes,totalUsuarios,totalUsuariosSIDAC)
			fila = completar(datosMes,fila,mes = mesesDict[mes])

	fila+=1
	ws['A{}'.format(fila)].value = 'Interrupción'
	ws['A{}'.format(fila)].fill = fillColor()
	ws['B{}'.format(fila)].value = 'Orden_Reposicion'
	ws['B{}'.format(fila)].fill = fillColor()
	ws['C{}'.format(fila)].value = 'Usuarios Afectados'
	ws['C{}'.format(fila)].fill = fillColor()
	ws['D{}'.format(fila)].value = 'Suma Min Corte'
	ws['D{}'.format(fila)].fill = fillColor()
	ws['E{}'.format(fila)].value = 'Inicio de la Int'
	ws['E{}'.format(fila)].fill = fillColor()
	ws['F{}'.format(fila)].value = 'Fin de la Int'
	ws['F{}'.format(fila)].fill = fillColor()
	ws['G{}'.format(fila)].value = 'Motivo_EPRE'
	ws['G{}'.format(fila)].fill = fillColor()
	fila+=1
	for corte in interrupciones:
		ws.cell(fila,1).value = corte[0]
		ws.cell(fila,2).value = corte[1]
		ws.cell(fila,3).value = corte[2]
		ws.cell(fila,4).value = int(corte[2])*(corte[4]-corte[3]).total_seconds()/60
		ws.cell(fila,5).value = corte[3]
		ws.cell(fila,6).value = corte[4]
		ws.cell(fila,7).value = corte[5]
		fila+=1

	wb.save('CCSPU y TCSPU/Indices CCSPU y TCSPU.xlsx')
	print('OK')