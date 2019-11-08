import zipfile
import pyodbc
import openpyxl
from rutas import dbHisto
from os import walk,mkdir
from os.path import isdir
from shutil import rmtree
from re import search

def getDir():
	global dir
	while True:
		dir = input('Carpeta:> ')
		if not isdir(r'{}'.format(dir)):
			print('No existe el directorio ingresado.')
			continue
		else: break
	if not dir: dir = getcwd()
		
def getFiles(folder):
	archivos = []
	for root, folders, files in walk(folder):
		archivos += files
	
	archivosR32 = [x for x in archivos if x.lower().endswith('.r32')]
	resultadosESJ = [x for x in archivos if search('(?i)Resultados (?:Re)?mediciones',x) and x.endswith('.xlsx')]
	resultadosESJxls = [x for x in archivos if search('(?i)Resultados (?:Re)?mediciones',x) and x.endswith('.xls')]
	if not resultadosESJ and resultadosESJxls:
		salir(mensaje='La extension del informe de Esj en esta carpeta es incorrecta. Tiene que ser tipo .xlsx')
	zips = [x for x in archivos if x.lower().endswith('.zip')]
	return archivosR32,resultadosESJ,zips

def extractZip(folder,zip):
	
	try: zip_ref.extractall(folder)
	except PermissionError:
		print('Alguno de los archivos esta siendo usado. Por favor cerralo y ejecutá el programa devuelta.')
		clean(folder)
		salir()
	
def pintarExcelEsj(folder,informe,diferencias):
	resultadosMedicionesWb = openpyxl.load_workbook('{}/{}'.format(folder,informe))
	resultadosMediciones = resultadosMedicionesWb['Rdo Mediciones Usuario-Centros']
	for row in resultadosMediciones:
		for par in diferencias:
			if par['sum'] == row[0].value and par['r32'] == row[16].value:
				for celda in row:
					celda.fill = openpyxl.styles.PatternFill(start_color='FFEEEE00',end_color='FFEEEE00',fill_type='solid')
				break
	resultadosMedicionesWb.save('{}/Analisis.xlsx'.format(folder))
	
def getListaArchivosR32(folder,informe):
	listadoR32 = []
	resultadosMedicionesWb = openpyxl.load_workbook('{}/{}'.format(folder,informe))
	resultadosMediciones = resultadosMedicionesWb['Rdo Mediciones Usuario-Centros']
	fila = 9
	while resultadosMediciones['A%d'%fila].value != None:
		suministro = resultadosMediciones['A%d'%fila].value
		archivo = resultadosMediciones['Q%d'%fila].value
		if archivo.endswith('.R32'): listadoR32.append({'sum':suministro,'r32':archivo})
		fila += 1
	resultadosMedicionesWb.close()
	return listadoR32
	
def getDataHisto(paresSumR32):
	conn = pyodbc.connect(r'Driver={Microsoft Access Driver (*.mdb, *.accdb)};DBQ=%s;' %dbHisto)
	cursor = conn.cursor()
	questionMarks = ','.join(('?' for _ in paresSumR32))
	query = 'SELECT Suministro,Archivo from Histo WHERE (Suministro IN ({}) AND Archivo IN ({}));'.format(questionMarks,questionMarks)
	cursor.execute(query,[x['sum'] for x in paresSumR32]+[x['r32'] for x in paresSumR32])
	lista = cursor.fetchall()
	lista = tuple(({'sum':x[0],'r32':x[1]} for x in lista))
	return lista
	
def salir(mensaje=''):
	print(mensaje)
	input('Enter para terminar...')
	exit()
	
def clean(folder):
	try: rmtree(folder)
	except:	print('No se pudo eliminar la carpeta temporal. Eliminala manualmente más tarde.')
	
def main():
	getDir()
	print()
	_,resultadosESJ,zips = getFiles(dir)
	for zip in zips:
		if search('(?i)Mediciones de Usuarios y Centros de',zip):
			comprimidoZip = zip
			tipoInforme = 'Mensual'
			break
	else:
		if len(zips )> 1:
			c = 1
			for z in zips:
				print(c,') ',z)
				c+=1
			choice = int(input('Numero de comprimido:> '))-1
			comprimidoZip = zips[choice]
		else: comprimidoZip = zips[0]
		tipoInforme = 'Diario'
	zip_ref = zipfile.ZipFile(dir+'/'+comprimidoZip, 'r')
	listaArchivosEnComprimido = zip_ref.namelist()
	archivosR32 = [x for x in listaArchivosEnComprimido if x.lower().endswith('.r32')]
	
	if tipoInforme == 'Diario':
		folder = dir+'/'+'Temp'
		try:mkdir(folder)
		except FileExistsError: pass
		excelEsj = [x for x in listaArchivosEnComprimido if x.startswith('Resultados Mediciones')]
		if excelEsj:
			excelEsj = excelEsj[0]
			zip_ref.extract(excelEsj,folder)
			paresSumR32 = getListaArchivosR32(folder,excelEsj)
		else: salir(mensaje='No se encontró el informe de ESJ en el comprimido')
			
		setArchivosR32 = set(archivosR32)
		setListadoR32 = set([x['r32'] for x in paresSumR32])
		if len(setArchivosR32) == len(setListadoR32):
			print('La cantidad de archivos enviados es igual a la cantidad de archivos informados.')
			mismaCantidad = True
		else:
			print('Existen diferencias en la cantidad de archivos enviados e informados.')
			mismaCantidad = False
		print()
		diferencias = setListadoR32.difference(setArchivosR32)
		if not diferencias:
			diferencias = setArchivosR32.difference(setListadoR32)
			if not diferencias:
				print('Estan todos los archivos informados en el informe de ESJ.')
				clean(folder)
				salir()
			else:
				print('Los siguientes archivos R32 fueron enviados pero no aparecen en el listado de ESJ:')
				for archivo in diferencias:
					print(archivo)
				if mismaCantidad:
					print('\nPuede que los archivos hayan sido enviados pero que difieran en el nombre.')
				clean(folder)
				salir()
		else:
			print('Los siguientes archivos R32 aparecen en el listado de ESJ pero no fueron enviados en el comprimido:')
			for archivo in diferencias:
				print(archivo)
			if mismaCantidad:
					print('\nPuede que los archivos hayan sido enviados pero que difieran en el nombre.')
			clean(folder)
			salir()
			
	elif tipoInforme == 'Mensual':
		excelEsj = resultadosESJ[0]
		paresSumR32 = getListaArchivosR32(dir,excelEsj)
		archivosProcesados = getDataHisto(paresSumR32)
		diferencias = tuple((x for x in paresSumR32 if x not in archivosProcesados))
		if len(diferencias): print('Hay {} mediciones que no estan en la tabla histo.'.format(len(diferencias)))
		else: print('Wohooo!!! Estan todas las mediciones en la tabla histo.')
		pintarExcelEsj(dir,excelEsj,diferencias)
		salir()
	
if __name__ == '__main__':
	main()