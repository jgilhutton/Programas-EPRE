import DataSuministro
import Mediciones
import CortesTresAños
import openpyxl
import pyodbc
import xlrd
from rutas import db,plantillaRotura,dbReclamosEpre
from re import search,sub
from time import strptime
from os import walk
# from string import ascii_uppercase
fila = 1

def getParesAnioMes(inicio,fin):
	pares = []
	if inicio.tm_year == fin.tm_year:
		for mes in range(inicio.tm_mon,fin.tm_mon+1):pares.append((inicio.tm_year,mes))
	else:
		if len(list(range(inicio.tm_year,fin.tm_year+1))) > 2:
			for año in range(inicio.tm_year,fin.tm_year+1)[1:-1]:
				for mes in range(1,13): pares.append((año,mes))
		for mes in range(inicio.tm_mon,13): pares.append((inicio.tm_year,mes))
		for mes in range(1,fin.tm_mon+1): pares.append((fin.tm_year,mes))
	return sorted(pares)
		

def getDataSuministrosSIDAC(suministro):
	query = 'SELECT * from suministros_sidac WHERE suministros_sidac.[CLIENTE] = ?;'
	connUsuarios = pyodbc.connect(r'Driver={Microsoft Access Driver (*.mdb, *.accdb)};DBQ=%s;' %db)
	cursorUsuarios = connUsuarios.cursor()
	cursorUsuarios.execute(query,(suministro,))
	datos = cursorUsuarios.fetchall()[0]
	return datos

def getReclamos(db,suministro=False,seta=False,distribuidor=False):
	try:connReclamos = pyodbc.connect(r'Driver={Microsoft Access Driver (*.mdb, *.accdb)};DBQ=%s;' %db)
	except: print('No se pudo conectar a la base de datos: {}'.format(db))
	cursorReclamos = connReclamos.cursor()
	tablas = [x[2] for x in cursorReclamos.tables()]
	tablaReclamos = list(filter(lambda x: search('(?i)reclamo',x),tablas))[0]
	if suministro:
		query = 'SELECT [RECLAMOS].* FROM [RECLAMOS] WHERE [RECLAMOS].[NumerodeSuministro] = ?;'
		query = query.replace('RECLAMOS',tablaReclamos)
		cursorReclamos.execute(query,(suministro,))
	elif seta:
		query = 'SELECT [RECLAMOS].* FROM [RECLAMOS] WHERE [RECLAMOS].[NumeroSeta] = ?;'
		query = query.replace('RECLAMOS',tablaReclamos)
		cursorReclamos.execute(query,(seta,))
	elif distribuidor:
		query = 'SELECT [RECLAMOS].* FROM [RECLAMOS] WHERE [RECLAMOS].[Distribuidor] = ?;'
		query = query.replace('RECLAMOS',tablaReclamos)
		cursorReclamos.execute(query,(distribuidor,))
	reclamos = cursorReclamos.fetchall()
	return reclamos
	
def getReclamosEpre(db,año,suministro):
	wb = xlrd.open_workbook(db)
	reclamosEPRE = wb.sheet_by_name(str(año))
	filas = reclamosEPRE.nrows
	reclamos = []
	for fila in range(0,filas):
		if reclamosEPRE.cell(fila,8).value == suministro:
			rec = [x.value for x in reclamosEPRE.row(fila)][1:]
			reclamos.append(rec)
	return reclamos
	
def getDistriET():
	query = 'SELECT distris.DISTRI,distris.NOMBRE AS DISTRIBUIDOR,centros.ET FROM distris INNER JOIN centros ON distris.CENTRO = centros.CODISE;'
	connUsuarios = pyodbc.connect(r'Driver={Microsoft Access Driver (*.mdb, *.accdb)};DBQ=%s;' %db)
	cursorUsuarios = connUsuarios.cursor()
	cursorUsuarios.execute(query)
	tabla = cursorUsuarios.fetchall()
	tupla = list(filter(lambda x: x[1] == dataSuministro['Distribuidor'],tabla))
	if tupla:
		codigoDistribuidor,distribuidor,et = tupla[0]
	distrisET = [y[0] for y in filter(lambda x: x[2] == et,tabla)]
	return codigoDistribuidor,distribuidor,et,distrisET
	
def getCortesDistribuidorET(db):
	connCortes = pyodbc.connect(r'Driver={Microsoft Access Driver (*.mdb, *.accdb)};DBQ=%s;' %db)
	cursorCortes = connCortes.cursor()
	query = 'SELECT Interrupcion,Orden_Reposicion,Inicio_EPRE,Fin_EPRE,Motivo_EPRE,Distribuidor FROM [Interrupciones_Reposiciones_Mensual] WHERE (Distribuidor LIKE ? AND Interrupcion LIKE ?);'
	cursorCortes.execute(query,(codigoDistribuidor,'MT%',))
	interrupcionesDistribuidor = cursorCortes.fetchall()
	query = 'SELECT [Interrupciones_Reposiciones_Mensual].Interrupcion,[Interrupciones_Reposiciones_Mensual].Orden_Reposicion,[Interrupciones_Reposiciones_Mensual].Inicio_EPRE,[Interrupciones_Reposiciones_Mensual].Fin_EPRE,[Interrupciones_Reposiciones_Mensual].Motivo_EPRE,[Interrupciones_Reposiciones_Mensual].Distribuidor FROM [Interrupciones_Reposiciones_Mensual] WHERE [Interrupciones_Reposiciones_Mensual].Distribuidor IN ({}) AND [Interrupciones_Reposiciones_Mensual].Interrupcion LIKE ?;'.format(','.join(['?' for _ in range(len(distrisET))]))
	cursorCortes.execute(query,distrisET+['MT%',])
	interrupcionesET = cursorCortes.fetchall()
	return interrupcionesDistribuidor,interrupcionesET

def setColor(plantilla,columnas,fila):
	for col in columnas:
		plantilla.cell(fila,col).fill = openpyxl.styles.PatternFill(start_color='FFBBBBBB',end_color='FFBBBBBB',fill_type='solid')
	return plantilla

def completar(plantilla,headers,datos,fila):
	for i,campo in enumerate(headers,start=1):
		plantilla.cell(fila,i).value = campo
	plantilla = setColor(plantilla,range(1,len(headers)+1),fila)
	fila+=1
	if datos:
		for dato in datos:
			for column,item in enumerate(dato,start=1):
				plantilla.cell(fila,column).value = item
			fila+=1
	fila+=2
	return plantilla,fila
	
def saveData(plantilla,fila):
	for column,key in enumerate(dataSuministro.keys(),start=1):
		plantilla.cell(fila,column).value = key
		plantilla.cell(fila+1,column).value = dataSuministro[key]
		plantilla = setColor(plantilla,range(1,len(dataSuministro.keys())+1),fila)
	fila+=2
	plantilla,fila = completar(plantilla,headersSIDAC,[dataSuministroSIDAC,],fila)
	
	plantilla['A{}'.format(fila)].value = 'Cortes Distribuidor'
	fila+=1
	plantilla,fila = completar(plantilla,headersCortes,cortesDis,fila)
	plantilla['A{}'.format(fila)].value = 'Cortes ET'
	fila+=1
	plantilla,fila = completar(plantilla,headersCortes,cortesET,fila)
	plantilla['A{}'.format(fila)].value = 'Cortes'
	fila+=1
	plantilla,fila = completar(plantilla,['Suministro','Interrupcion','Orden_Reposicion','Inicio','Final','Id','Motivo_EPRE'],cortes,fila)
	plantilla['A{}'.format(fila)].value = 'Mediciones en el suministro'
	fila+=1
	plantilla,fila = completar(plantilla,headersMediciones,mediciones,fila)
	plantilla['A{}'.format(fila)].value = 'Reclamos a ESJ'
	fila+=1
	plantilla,fila = completar(plantilla,headersReclamos,reclamosUsuario,fila)
	plantilla['A{}'.format(fila)].value = 'Reclamos de la SETA'
	fila+=1
	plantilla,fila = completar(plantilla,headersReclamos,reclamosSeta,fila)
	plantilla['A{}'.format(fila)].value = 'Reclamos del distribuidor'
	fila+=1
	plantilla,fila = completar(plantilla,headersReclamos,reclamosDistribuidor,fila)
	plantilla['A{}'.format(fila)].value = 'Reclamos en el EPRE'
	fila+=1
	plantilla,fila = completar(plantilla,headersEPRE,reclamosEpre,fila)
	return plantilla

def getPathDbReclamos(año,mes):
	semestre = '01' if mes in [1,2,3,4,5,6] else '02'
	año,mes = str(año),str(mes)
	mesLargo = {'1':'Enero','2':'Febrero','3':'Marzo','4':'Abril','5':'Mayo','6':'Junio','7':'Julio','8':'Agosto','9':'Septiembre','10':'Octubre','11':'Noviembre','12':'Diciembre'}[mes]
	for root,dirs,files in walk('//Alfredo/Servidor/ACS/SERVICIO TECNICO/ESJ/2º ETAPA/{}{}/'.format(año[2:],semestre)):
		if mesLargo.lower() in root.lower() and 'Canal EPRE Etapa 2.accdb' in files:
			path = root.replace('\\','/')+'/Canal EPRE Etapa 2.accdb'
			break
	else:
		print('No se encontro la base de datos para {} {}{}'.format(mesLargo,año[2:],semestre))
		return None
	return path
	
if __name__ == '__main__':
	headersCortes = ['Interrupcion','Orden_Reposicion','Inicio_EPRE','Final_EPRE','Motivo_EPRE','Distribuidor']
	headersEPRE = ['Fecha', 'HORA', 'Motivo del Reclamo', 'Nombre y Apellido', 'Teléfono', 'Domicilio ', 'Dpto', 'Nº Suministro', 'Reclamos anteriores si los tuviere', 'Nº RECLAMO E.S.J. ', 'Atendió en EPRE', 'Tareas de normalización realizadas y/o previstas x Energía San Juan', 'Comunicación EPRE posterior con el usuario', 'Tipo de Reclamo']
	headersReclamos = ['NumerodeOrden','NumerodeReclamo','FechayHoradeReclamo','NumerodeSuministro','NombreUsuario','Domicilio','Departamento','VillaoBarrio','NumeroSeta','MotivoReclamo','CodigoTrabajo','TrabajoRealizado','FechaHoraInicioAtencion','FechaHoraFinAtencion','FechaHoraLlamado','Recepcionista','Descripción Falla','nombre_contratista1','Fecha Pasado a 1','nombre_contratista2','Fecha derivado a 2','OAC','Distribuidor','Interrupcion']
	headersSIDAC = ['SETA','CLIENTE','NOMBRE','CALLE','NUMERO','DPTO','LOCALIDAD','CATEGORIA','ESTADO','SITUACION','MEDIDOR','MARCA','REMESADO','UNIDAD_LECTURA','NRO_SECUENCIA','SAC_CONEXION','TARIFA']
	headersMediciones = ['FECHA-COL','FECHA- RET','RESULTADO','SETA','PTO-MED']

	print('Suministro:> XXXXXXXXXXX\r',end='Suministro:> ')
	suministro = input()
	print('Inicio periodo de análisis')
	print('Fecha:> dd/mm/aa\r',end='Fecha:> ')
	fechaInicio = strptime(input(),'%d/%m/%y')
	print('Fin periodo de análisis')
	print('Fecha:> dd/mm/aa\r',end='Fecha:> ')
	fechaFin = strptime(input(),'%d/%m/%y')
		
	print('Obteniendo informacion del suministro... \r',end='Obteniendo informacion del suministro... ')
	dataSuministro = DataSuministro.query(suministro)
	dataSuministroSIDAC = getDataSuministrosSIDAC(suministro)
	codigoDistribuidor,distribuidor,et,distrisET = getDistriET()
	print('OK')

	print('Obteniendo cortes del distribuidor y et... \r',end='Obteniendo cortes del distribuidor y et... ')
	cortesDis,cortesET,reclamosUsuario,reclamosSeta,reclamosDistribuidor,reclamosEpre = [],[],[],[],[],[]
	paresAñoMes =  getParesAnioMes(fechaInicio,fechaFin)
	for año,mes in paresAñoMes:
		dbReclamos = getPathDbReclamos(año,mes)
		cortesDisTemp,cortesETTemp = getCortesDistribuidorET(dbReclamos)
		cortesDis += cortesDisTemp
		cortesET += cortesETTemp
		reclamosUsuario += getReclamos(dbReclamos,suministro=suministro)
		reclamosSeta += getReclamos(dbReclamos,seta=dataSuministro['Seta'])
		reclamosDistribuidor += getReclamos(dbReclamos,distribuidor=dataSuministro['Distribuidor'])
	print('OK')

	print('Obteniendo reclamos a EPRE... \r',end='Obteniendo reclamos a EPRE... ')
	for año in set([x[0] for x in paresAñoMes]):
		dbTemp = sub('YYYY',str(año),dbReclamosEpre)
		reclamosEpre += getReclamosEpre(dbTemp,año,suministro)
	print('OK')

	print('Obteniendo mediciones... \r',end='Obteniendo mediciones... ')
	mediciones = Mediciones.forImport([suministro,])
	print('OK')

	print('Obteniendo cortes... \r',end='Obteniendo cortes... ')
	cortes = CortesTresAños.forImport(suministro,False,años=1)[1]
	print('OK')

	libroExcel = openpyxl.load_workbook(plantillaRotura)
	plantilla = libroExcel['Informacion']
	plantilla = saveData(plantilla,fila)
	libroExcel.save('Data para Rotura.xlsx')