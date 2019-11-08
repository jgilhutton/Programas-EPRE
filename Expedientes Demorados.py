import xlrd
import openpyxl
from datetime import datetime,timedelta
from rutas import listadoExpedientes, plantillaDemorados
from re import search

regex = '(?i)(?<!\w)AL(?!=\w)|IGG|PGG|N(?:[U,D]|ota),pase,GG'

print('Minima demora: 5  dias\r',end='Minima demora: ')
maximoDias = input()
maximoDias = int(maximoDias) if maximoDias else 5
print('Buscar 90 dias hacia atras\r',end='Buscar ')
minimoDias = input()
minimoDias = int(minimoDias) if minimoDias else 90
epoch = datetime(1900,1,1,0,0,0,0)
hoy = datetime.now()

wb = xlrd.open_workbook(listadoExpedientes)
ws = wb.sheet_by_name('EXP. EN PODER ') # si.... hay un espacio al final

demorados = []

for r in range(3,ws.nrows):
	row = ws.row(r)[::-1]
	
	pgg = False
	cellNo = -1
	for cell in row:
		cellNo += 1
		if pgg:
			try: fechaInt = int(cell.value)
			except ValueError: break
			fechaSalida = epoch + timedelta(days=fechaInt)
			diasDesdeSalida = (hoy - fechaSalida).days
			if minimoDias > diasDesdeSalida > maximoDias:
				demorados.append([str(row[-1].value).strip(),fechaSalida.strftime('%d/%m/%Y'),str(pgg).strip()])
				break
			 
		if not cell.value or cellNo == 1: continue
		else:
			if search(regex,str(cell.value).lower()):
				pgg = cell.value
				continue
			else:
				break
	
wbOut = openpyxl.load_workbook(plantillaDemorados)
wsOut = wbOut['Hoja1']
for index,row in enumerate(demorados,start=1):
	wsOut.cell(index,1).value = str(row[0]).strip()
	wsOut.cell(index,2).value = row[1]
	wsOut.cell(index,3).value = str(row[2]).strip()
wbOut.save('Demorados.xlsx')
	